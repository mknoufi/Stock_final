from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest

from backend.db.indexes import INDEXES
from backend.db.migrations import MigrationManager
from backend.services.advanced_report_service import (
    AdvancedReportService,
    ColumnConfig,
    ReportConfig,
    ReportFilters,
)
from backend.services.scheduled_export_service import ScheduledExportService


class _AggregateCursor:
    def __init__(self, rows):
        self._rows = list(rows)

    async def to_list(self, length):
        return self._rows[:length]


@pytest.mark.asyncio
async def test_verified_items_report_uses_alias_aware_match_and_projection():
    db = MagicMock()
    db.count_lines.count_documents = AsyncMock(side_effect=[3, 1])
    pipelines = []

    def _aggregate(pipeline):
        pipelines.append(pipeline)
        if len(pipelines) == 1:
            return _AggregateCursor([])
        return _AggregateCursor([{"_id": None, "total_items": 0}])

    db.count_lines.aggregate.side_effect = _aggregate

    service = AdvancedReportService(db)
    await service.generate_verified_items_report(
        ReportConfig(
            report_type="verified_items",
            filters=ReportFilters(category="Electronics", floor="F1", rack_id="R-7"),
        )
    )

    match_stage = pipelines[0][0]["$match"]
    assert {"$or": [{"category": "Electronics"}, {"category_correction": "Electronics"}, {"category_erp": "Electronics"}]} in match_stage["$and"]
    assert {"$or": [{"floor": "F1"}, {"floor_no": "F1"}]} in match_stage["$and"]
    assert {"$or": [{"rack_id": "R-7"}, {"rack_no": "R-7"}]} in match_stage["$and"]

    project_stage = pipelines[0][-1]["$project"]
    assert project_stage["category"] == {
        "$ifNull": ["$category_correction", {"$ifNull": ["$category_erp", "$category"]}]
    }
    assert project_stage["floor"] == {"$ifNull": ["$floor", "$floor_no"]}
    assert project_stage["rack_id"] == {"$ifNull": ["$rack_id", "$rack_no"]}
    assert project_stage["stock_qty"] == {"$ifNull": ["$stock_qty", "$erp_qty"]}
    assert project_stage["notes"] == {
        "$ifNull": ["$notes", {"$ifNull": ["$remark", "$variance_note"]}]
    }


@pytest.mark.asyncio
async def test_scheduled_count_line_export_uses_counted_at_and_erp_qty():
    db = MagicMock()
    service = ScheduledExportService(db)

    sorted_cursor = AsyncMock()
    sorted_cursor.to_list.return_value = [
        {
            "_id": "mongo-1",
            "id": "line-1",
            "session_id": "session-1",
            "item_code": "ITM-1",
            "item_name": "Mixer",
            "barcode": "123456",
            "erp_qty": 8,
            "counted_qty": 6,
            "variance": -2,
            "variance_reason": "damaged",
            "counted_by": "staff1",
            "counted_at": "2026-03-22T10:00:00",
        }
    ]
    db.count_lines.find.return_value.sort.return_value = sorted_cursor

    rows = await service._export_count_lines({"session_id": "session-1"})

    db.count_lines.find.assert_called_once_with({"session_id": "session-1"})
    db.count_lines.find.return_value.sort.assert_called_once_with("counted_at", -1)
    assert rows == [
        {
            "line_id": "line-1",
            "session_id": "session-1",
            "item_code": "ITM-1",
            "item_name": "Mixer",
            "barcode": "123456",
            "erp_qty": 8,
            "counted_qty": 6,
            "variance": -2,
            "variance_reason": "damaged",
            "counted_by": "staff1",
            "counted_at": "2026-03-22T10:00:00",
        }
    ]


@pytest.mark.asyncio
async def test_migration_manager_adds_sparse_unique_idempotency_index():
    db = MagicMock()
    manager = MigrationManager(db)
    manager._create_index_safe = AsyncMock()

    await manager._ensure_count_lines_indexes()

    manager._create_index_safe.assert_any_call(
        db.count_lines,
        "idempotency_key",
        unique=True,
        sparse=True,
        name="count_lines.idempotency_key",
    )


def test_optimized_count_line_indexes_match_runtime_fields():
    count_line_indexes = {options["name"]: (fields, options) for fields, options in INDEXES["count_lines"]}

    assert count_line_indexes["idx_session_counts"][0] == [("session_id", 1), ("counted_at", -1)]
    assert count_line_indexes["idx_rack_counts"][0] == [("rack_no", 1), ("session_id", 1)]

    idempotency_fields, idempotency_options = count_line_indexes["idx_count_line_idempotency"]
    assert idempotency_fields == [("idempotency_key", 1)]
    assert idempotency_options["unique"] is True
    assert idempotency_options["sparse"] is True


@pytest.mark.asyncio
async def test_advanced_report_csv_export_supports_erpnext_import_headers():
    service = AdvancedReportService(MagicMock())
    columns = [
        ColumnConfig(field="item_code", label="Item Code"),
        ColumnConfig(field="counted_qty", label="Counted Qty"),
    ]

    csv_content = await service.export_to_csv(
        [{"item_code": "ITM-1", "counted_qty": 3}],
        columns,
        erpnext_import=True,
    )

    assert csv_content.splitlines()[0] == "ID,item_code,counted_qty"
    assert csv_content.splitlines()[1] == ",ITM-1,3"


@pytest.mark.asyncio
async def test_advanced_report_xlsx_export_supports_erpnext_import_headers():
    service = AdvancedReportService(MagicMock())
    columns = [
        ColumnConfig(field="item_code", label="Item Code"),
        ColumnConfig(field="counted_qty", label="Counted Qty"),
    ]

    workbook_bytes = await service.export_to_xlsx(
        [{"item_code": "ITM-1", "counted_qty": 3}],
        columns,
        erpnext_import=True,
    )
    workbook = openpyxl.load_workbook(filename=BytesIO(workbook_bytes))
    sheet = workbook.active

    assert sheet.cell(row=1, column=1).value == "ID"
    assert sheet.cell(row=1, column=2).value == "item_code"
    assert sheet.cell(row=1, column=3).value == "counted_qty"
    assert sheet.cell(row=2, column=1).value in ("", None)
    assert sheet.cell(row=2, column=2).value == "ITM-1"
    assert sheet.cell(row=2, column=3).value == 3
