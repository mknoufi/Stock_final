from unittest.mock import AsyncMock, Mock

from io import BytesIO
import pytest
from fastapi import HTTPException
from datetime import datetime, timezone
import openpyxl

from backend.api.item_verification_api import (
    ItemUpdateRequest,
    VerificationRequest,
    _build_erpnext_item_export_row,
    _build_erpnext_variance_export_row,
    build_item_filter_query,
    init_verification_api,
    sync_items_for_offline_cache,
    update_item_master,
    verify_item,
)


@pytest.fixture(autouse=True)
def setup_mocks():
    mock_db = AsyncMock()
    mock_cache = AsyncMock()
    init_verification_api(mock_db, mock_cache)
    return mock_db, mock_cache


@pytest.mark.asyncio
async def test_update_item_master_success(setup_mocks):
    mock_db, mock_cache = setup_mocks

    item_code = "CODE123"
    existing_item = {"item_code": item_code, "barcode": "510001", "mrp": 100.0}
    mock_db.erp_items.find_one.return_value = existing_item

    request = ItemUpdateRequest(mrp=150.0, sales_price=120.0, category="New Cat")
    current_user = {"username": "testuser"}

    response = await update_item_master(item_code, request, current_user)

    assert response["success"] is True

    # Verify DB update
    mock_db.erp_items.update_one.assert_called_once()
    call_args = mock_db.erp_items.update_one.call_args
    assert call_args[0][0]["item_code"] == item_code
    update_doc = call_args[0][1]["$set"]
    assert update_doc["mrp"] == 150.0
    assert update_doc["sales_price"] == 120.0
    assert update_doc["category"] == "New Cat"
    assert update_doc["last_updated_by"] == "testuser"

    # Verify cache invalidation
    assert mock_cache.delete_async.call_count >= 1

    # Verify audit log
    mock_db.audit_logs.insert_one.assert_called_once()


@pytest.mark.asyncio
async def test_update_item_master_not_found(setup_mocks):
    mock_db, _ = setup_mocks
    mock_db.erp_items.find_one.return_value = None

    request = ItemUpdateRequest(mrp=150.0)
    current_user = {"username": "testuser"}

    with pytest.raises(HTTPException) as exc:
        await update_item_master("NONEXISTENT", request, current_user)

    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_verify_item_success(setup_mocks):
    mock_db, mock_cache = setup_mocks

    item_code = "CODE123"
    existing_item = {
        "item_code": item_code,
        "stock_qty": 10.0,
        "floor": "Old Floor",
        "barcode": "510001",
    }
    # First find_one returns existing item, second find_one returns updated item
    updated_item = existing_item.copy()
    updated_item["_id"] = "mock_id"
    mock_db.erp_items.find_one.side_effect = [existing_item, updated_item]

    request = VerificationRequest(
        verified=True, verified_qty=8.0, damaged_qty=1.0, floor="New Floor", notes="Test notes"
    )
    current_user = {"username": "testuser"}

    response = await verify_item(item_code, request, current_user)

    assert response["success"] is True
    assert response["variance"] == -1.0  # 8 + 1 - 10 = -1

    # Verify DB update
    mock_db.erp_items.update_one.assert_called_once()
    call_args = mock_db.erp_items.update_one.call_args
    assert call_args[0][0]["item_code"] == item_code
    update_doc = call_args[0][1]["$set"]
    assert update_doc["verified"] is True
    assert update_doc["verified_qty"] == 8.0
    assert update_doc["variance"] == -1.0
    assert update_doc["verified_floor"] == "New Floor"

    # Verify cache invalidation
    assert mock_cache.delete_async.call_count >= 1

    # Verify logs
    mock_db.verification_logs.insert_one.assert_called_once()
    mock_db.item_variances.insert_one.assert_called_once()  # Variance is not 0


def test_build_item_filter_query():
    # Test basic filters
    query = build_item_filter_query(category="Cat1", verified=True, search="Test")

    assert query["category"] == {"$regex": "Cat1", "$options": "i"}
    assert query["verified"] is True
    assert "$or" in query
    assert len(query["$or"]) == 3

    # Test empty filters
    query = build_item_filter_query()
    assert query == {}


def test_build_erpnext_item_export_row_includes_blank_id():
    row = _build_erpnext_item_export_row(
        {
            "item_code": "ITM-1",
            "item_name": "Mixer",
            "barcode": "123456",
            "stock_qty": 8,
            "verified": True,
        }
    )

    assert row["ID"] == ""
    assert row["item_code"] == "ITM-1"
    assert row["item_name"] == "Mixer"
    assert row["barcode"] == "123456"
    assert row["stock_qty"] == 8
    assert row["verified"] == "Yes"


def test_build_erpnext_variance_export_row_includes_blank_id():
    row = _build_erpnext_variance_export_row(
        {
            "item_code": "ITM-9",
            "item_name": "Kettle",
            "system_qty": 5,
            "verified_qty": 3,
            "variance": -2,
            "count_line_id": "line-22",
        }
    )

    assert row["ID"] == ""
    assert row["item_code"] == "ITM-9"
    assert row["system_qty"] == 5
    assert row["verified_qty"] == 3
    assert row["variance"] == -2
    assert row["count_line_id"] == "line-22"


@pytest.mark.asyncio
async def test_sync_items_for_offline_cache_with_since(setup_mocks):
    mock_db, _ = setup_mocks

    since = datetime(2025, 1, 1, tzinfo=timezone.utc).replace(tzinfo=None)

    cursor = Mock()
    cursor.sort.return_value = cursor
    cursor.limit.return_value = cursor
    cursor.to_list = AsyncMock(
        return_value=[
            {
                "barcode": "510001",
                "item_code": "CODE123",
                "item_name": "Test Item",
                "category": "Cat",
                "verified": True,
                "verified_at": since,
                "last_scanned_at": since,
            }
        ]
    )
    # Motor's collection.find() is sync (returns a cursor), so use a normal Mock.
    mock_db.erp_items.find = Mock(return_value=cursor)

    response = await sync_items_for_offline_cache(
        since=since, limit=10, current_user={"username": "u"}
    )

    assert response["success"] is True
    assert response["count"] == 1
    assert len(response["items"]) == 1
    assert response["items"][0]["barcode"] == "510001"
    assert response["items"][0]["verified_at"] == since.isoformat()
    assert response["items"][0]["last_scanned_at"] == since.isoformat()

    # Ensure the query includes a since filter for change timestamps.
    find_args, _find_kwargs = mock_db.erp_items.find.call_args
    assert "barcode" in find_args[0]
    assert "$or" in find_args[0]


@pytest.mark.asyncio
async def test_sync_items_for_offline_cache_without_since(setup_mocks):
    mock_db, _ = setup_mocks

    cursor = Mock()
    cursor.sort.return_value = cursor
    cursor.limit.return_value = cursor
    cursor.to_list = AsyncMock(return_value=[])
    mock_db.erp_items.find = Mock(return_value=cursor)

    response = await sync_items_for_offline_cache(
        since=None, limit=5, current_user={"username": "u"}
    )

    assert response["success"] is True
    assert response["count"] == 0
    assert response["items"] == []

    find_args, _find_kwargs = mock_db.erp_items.find.call_args
    assert find_args[0] == {"barcode": {"$exists": True, "$ne": ""}}


@pytest.mark.asyncio
async def test_export_items_csv_route_returns_erpnext_headers(
    async_client,
    authenticated_headers,
    test_db,
):
    await test_db.erp_items.insert_one(
        {
            "item_code": "ITM-1",
            "item_name": "Mixer",
            "barcode": "123456",
            "stock_qty": 8,
            "verified": True,
        }
    )

    response = await async_client.get(
        "/api/v2/erp/items/export/csv",
        headers=authenticated_headers,
    )

    assert response.status_code == 200
    assert "items_erpnext_import_" in response.headers["content-disposition"]
    first_line = response.text.splitlines()[0]
    assert first_line.startswith("ID,item_code,item_name,barcode")


@pytest.mark.asyncio
async def test_export_items_xlsx_route_returns_erpnext_headers(
    async_client,
    authenticated_headers,
    test_db,
):
    await test_db.erp_items.insert_one(
        {
            "item_code": "ITM-2",
            "item_name": "Kettle",
            "barcode": "654321",
            "stock_qty": 5,
            "verified": False,
        }
    )

    response = await async_client.get(
        "/api/v2/erp/items/export/xlsx",
        headers=authenticated_headers,
    )

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "ID"
    assert sheet.cell(row=1, column=2).value == "item_code"
    assert sheet.cell(row=2, column=2).value == "ITM-2"


@pytest.mark.asyncio
async def test_export_variances_csv_route_returns_erpnext_headers(
    async_client,
    authenticated_headers,
    test_db,
):
    await test_db.item_variances.insert_one(
        {
            "item_code": "ITM-3",
            "item_name": "Toaster",
            "system_qty": 9,
            "verified_qty": 7,
            "variance": -2,
            "verified_by": "staff1",
            "verified_at": datetime.now(timezone.utc).replace(tzinfo=None),
        }
    )

    response = await async_client.get(
        "/api/v2/erp/items/variances/export/csv",
        headers=authenticated_headers,
    )

    assert response.status_code == 200
    assert "variances_erpnext_import_" in response.headers["content-disposition"]
    assert response.text.splitlines()[0].startswith("ID,item_code,item_name,system_qty")


@pytest.mark.asyncio
async def test_export_variances_xlsx_route_returns_erpnext_headers(
    async_client,
    authenticated_headers,
    test_db,
):
    await test_db.item_variances.insert_one(
        {
            "item_code": "ITM-4",
            "item_name": "Grinder",
            "system_qty": 4,
            "verified_qty": 6,
            "variance": 2,
            "verified_by": "staff1",
            "verified_at": datetime.now(timezone.utc).replace(tzinfo=None),
        }
    )

    response = await async_client.get(
        "/api/v2/erp/items/variances/export/xlsx",
        headers=authenticated_headers,
    )

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "ID"
    assert sheet.cell(row=1, column=2).value == "item_code"
    assert sheet.cell(row=2, column=2).value == "ITM-4"
